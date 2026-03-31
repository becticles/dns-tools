#!/usr/bin/env python3

import csv
import os
import argparse
from pathlib import Path
from collections import defaultdict
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ============================================
# CONSTANTS
# ============================================

SKIP_SHEET_NAMES = {"Appliances"}

EXTERNAL_IP_COLUMN = 2   # Column B
INTERNAL_IP_COLUMN = 3   # Column C
START_ROW = 2

OUTPUT_INT_COUNT_COLUMN = 10  # J
OUTPUT_INT_DNS_COLUMN   = 11  # K
OUTPUT_EXT_COUNT_COLUMN = 12  # L
OUTPUT_EXT_DNS_COLUMN   = 13  # M

LEGACY_SHEET_NAME = "Legacy VIP Candidates"

LIKELY_VIP_FILENAMES = [
    "f5_VIPs.xlsx",
    "F5_VIPs.xlsx",
    "vip_inventory.xlsx",
]

# ============================================
# HELPERS
# ============================================

def clean_text(val):
    if val is None:
        return ""
    return str(val).strip().strip('"').strip()

def clean_fqdn(val):
    return clean_text(val).rstrip(".").lower()

def clean_ip(val):
    return clean_text(val)

def is_ipv4(value):
    value = clean_ip(value)
    parts = value.split(".")
    if len(parts) != 4:
        return False
    try:
        return all(0 <= int(p) <= 255 for p in parts)
    except ValueError:
        return False

def build_fqdn(hostname, zone):
    hostname = clean_text(hostname)
    zone = clean_text(zone)

    if hostname == "@":
        return clean_fqdn(zone)

    if hostname and zone:
        return clean_fqdn(f"{hostname}.{zone}")

    if hostname:
        return clean_fqdn(hostname)

    if zone:
        return clean_fqdn(zone)

    return ""

def ptr_zone_to_ipv4(hostname, zone):
    hostname = clean_text(hostname)
    zone = clean_text(zone).lower().rstrip(".")

    if not hostname or not zone.endswith(".in-addr.arpa"):
        return ""

    zone_prefix = zone[:-len(".in-addr.arpa")]
    zone_parts = zone_prefix.split(".")
    zone_parts.reverse()

    host_parts = hostname.split(".")
    host_parts.reverse()

    ip_parts = zone_parts + host_parts

    if len(ip_parts) != 4:
        return ""

    ip = ".".join(ip_parts)
    return ip if is_ipv4(ip) else ""

def recreate_sheet(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    return wb.create_sheet(title=sheet_name)

def autosize_columns(ws, min_width=12, max_width=80):
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(
            min_width, min(max_len + 2, max_width)
        )

# ============================================
# CONFIG / PATHS
# ============================================

def parse_args():
    parser = argparse.ArgumentParser(
        description="Reconcile F5 VIP inventory with DNS exports."
    )
    parser.add_argument(
        "--dns",
        help="Directory containing DNS CSV exports. Overrides DNS_FOLDER env var."
    )
    parser.add_argument(
        "--vip",
        help="VIP workbook path. Overrides VIP_WORKBOOK env var."
    )
    parser.add_argument(
        "--output",
        help="Output directory. Overrides OUTPUT_FOLDER env var."
    )
    return parser.parse_args()

def find_vip_workbook_local():
    print("Searching for VIP workbook in current directory...")

    direct_candidates = [Path(name) for name in LIKELY_VIP_FILENAMES]

    for candidate in direct_candidates:
        if candidate.exists():
            resolved = candidate.resolve()
            print("Using VIP workbook:", resolved)
            return resolved

    raise FileNotFoundError(
        "Could not locate VIP workbook automatically in the current directory. "
        "Use --vip /path/to/workbook.xlsx or set VIP_WORKBOOK."
    )

def resolve_paths(args):
    dns_env = os.getenv("DNS_FOLDER")
    vip_env = os.getenv("VIP_WORKBOOK")
    output_env = os.getenv("OUTPUT_FOLDER")

    dns_folder = Path(args.dns) if args.dns else Path(dns_env) if dns_env else Path("./dns_exports")
    output_folder = Path(args.output) if args.output else Path(output_env) if output_env else Path("./output")

    if args.vip:
        vip_workbook = Path(args.vip)
    elif vip_env:
        vip_workbook = Path(vip_env)
    else:
        vip_workbook = find_vip_workbook_local()

    output_folder.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    output_workbook = output_folder / f"vip_dns_report_{timestamp}.xlsx"

    return dns_folder, vip_workbook, output_workbook

def print_source_diagnostics(vip_workbook, dns_folder, output_workbook):
    print("DNS folder:", dns_folder)
    print("Source workbook:", vip_workbook)
    print("Output workbook:", output_workbook)

    if not vip_workbook.exists():
        raise FileNotFoundError(f"Source workbook not found: {vip_workbook}")

    if not dns_folder.exists():
        raise FileNotFoundError(f"DNS folder not found: {dns_folder}")

    stat = vip_workbook.stat()
    modified = datetime.fromtimestamp(stat.st_mtime)

    print("Source exists: True")
    print("Source size bytes:", stat.st_size)
    print("Source modified:", modified)

# ============================================
# DNS LOADER
# ============================================

def load_dns_records(folder):
    ip_to_dns = defaultdict(set)

    csv_files = sorted(folder.glob("*.csv"))
    if not csv_files:
        raise RuntimeError(f"No DNS CSV files found in {folder}")

    print("DNS files detected:")
    for csv_file in csv_files:
        print("  ", csv_file.name)

    for csv_file in csv_files:
        print("Reading", csv_file.name)

        with open(csv_file, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)

            for row in reader:
                rtype = clean_text(row.get("RecordType", "")).upper()
                hostname = row.get("HostName", "")
                zone = row.get("Zone", "")
                rdata = row.get("RecordData", "")

                if rtype == "A":
                    fqdn = build_fqdn(hostname, zone)
                    ip = clean_ip(rdata)

                    if fqdn and is_ipv4(ip):
                        ip_to_dns[ip].add(fqdn)

                elif rtype == "PTR":
                    fqdn = clean_fqdn(rdata)
                    ip = ptr_zone_to_ipv4(hostname, zone)

                    if fqdn and is_ipv4(ip):
                        ip_to_dns[ip].add(fqdn)

    if not ip_to_dns:
        raise RuntimeError(
            f"No DNS records loaded from {folder}. Check your DNS export files."
        )

    return ip_to_dns

# ============================================
# WORKBOOK PROCESSING
# ============================================

def update_workbook(vip_workbook, output_workbook, ip_to_dns):
    print("Opening workbook...")
    wb = load_workbook(vip_workbook)
    print("Loaded sheets:", ", ".join(wb.sheetnames))

    legacy_rows = []

    for ws in wb.worksheets:
        if ws.title in SKIP_SHEET_NAMES:
            print("Skipping sheet:", ws.title)
            continue

        print("Processing sheet:", ws.title)

        ws.cell(1, OUTPUT_INT_COUNT_COLUMN, "Matched DNS Count")
        ws.cell(1, OUTPUT_INT_DNS_COLUMN, "Matched DNS")
        ws.cell(1, OUTPUT_EXT_COUNT_COLUMN, "External DNS Count")
        ws.cell(1, OUTPUT_EXT_DNS_COLUMN, "External DNS")

        for row in range(START_ROW, ws.max_row + 1):
            if row % 500 == 0:
                print(f"  {ws.title}: row {row}/{ws.max_row}")

            internal_ip = clean_ip(ws.cell(row, INTERNAL_IP_COLUMN).value)
            external_ip = clean_ip(ws.cell(row, EXTERNAL_IP_COLUMN).value)

            internal_matches = []
            external_matches = []

            if is_ipv4(internal_ip):
                internal_matches = sorted(ip_to_dns.get(internal_ip, set()))

            if is_ipv4(external_ip):
                external_matches = sorted(ip_to_dns.get(external_ip, set()))

            ws.cell(row, OUTPUT_INT_COUNT_COLUMN, len(internal_matches))
            ws.cell(row, OUTPUT_INT_DNS_COLUMN, ", ".join(internal_matches))
            ws.cell(row, OUTPUT_EXT_COUNT_COLUMN, len(external_matches))
            ws.cell(row, OUTPUT_EXT_DNS_COLUMN, ", ".join(external_matches))

            if (
                (is_ipv4(internal_ip) or is_ipv4(external_ip))
                and len(internal_matches) == 0
                and len(external_matches) == 0
            ):
                legacy_rows.append([
                    ws.title,
                    row,
                    clean_text(ws.cell(row, 1).value),
                    external_ip,
                    internal_ip
                ])

    legacy_ws = recreate_sheet(wb, LEGACY_SHEET_NAME)

    legacy_ws.append([
        "Sheet",
        "Row",
        "VIP Name",
        "External IP",
        "Internal IP"
    ])

    for r in legacy_rows:
        legacy_ws.append(r)

    autosize_columns(legacy_ws)

    print("Saving:", output_workbook)
    wb.save(output_workbook)
    print("Legacy VIP candidates:", len(legacy_rows))

# ============================================
# MAIN
# ============================================

def main():
    args = parse_args()
    dns_folder, vip_workbook, output_workbook = resolve_paths(args)

    print_source_diagnostics(vip_workbook, dns_folder, output_workbook)

    print("Loading DNS records...")
    ip_to_dns = load_dns_records(dns_folder)
    print("Unique IPs found:", len(ip_to_dns))

    update_workbook(vip_workbook, output_workbook, ip_to_dns)

    print("Done.")

if __name__ == "__main__":
    main()